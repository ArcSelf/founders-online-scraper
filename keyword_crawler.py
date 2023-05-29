import requests
from bs4 import BeautifulSoup
import spacy
import random
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

# Load Spacy model and add sentencizer to the pipeline
nlp = spacy.load("en_core_web_sm")
nlp.add_pipe('sentencizer')

# Base URL of Thomas Jefferson's letters
base_url = "https://founders.archives.gov/?q=Project%3A%22Jefferson%20Papers%22&s=1511311111&r="

# Number of pages to process
num_pages = 100

# Prepare the keywords list
keywords = [
    "military", "servicemen", "service men", "virtue", "soldier", "army", "navy", "war",
    "battle", "valor", "honor", "patriotism", "nation", "freedom", "duty", "sacrifice",
    "courage", "bravery", "constitution", "declaration", "independence", "liberty",
    "democracy", "republic", "justice", "rights", "right", "federal", "union", "law",
    "government", "congress", "president", "states", "territory", "foreign", "treaty",
    "peace", "revolution", "colonies", "france", "britain", "spain", "trade", "commerce",
    "slavery", "agriculture", "education", "religion", "science", "enlightenment", "reason",
    "history", "philosophy", "economy", "industry", "culture", "morality", "ethics",
    "humanity", "citizenship", "equality", "diversity", "tradition", "legacy", "sovereignty",
    "leadership", "policy", "legislation", "diplomacy", "alliance", "conflict", "resolution",
    "prosperity", "crisis", "reform", "movement", "protest", "resistance", "innovation",
    "exploration", "discovery", "expansion", "frontier", "settlement", "pioneer", "immigration",
    "stoic", "ethics", "metaphysics", "foundation", "revolution"
]

# Convert the keywords to their lemma forms and to a set
keywords = set(nlp(keyword)[0].lemma_ for keyword in keywords)

# Initialize reference count
reference_count = 0

# Workbook path
workbook_path = "references.xlsx"

# Try to load the workbook, create a new one if it does not exist
try:
    workbook = load_workbook(workbook_path)
except FileNotFoundError:
    workbook = Workbook()

# Select the worksheet or create a new one if it does not exist
sheet = workbook["Thomas Jefferson"] if "Thomas Jefferson" in workbook.sheetnames else workbook.create_sheet("Thomas Jefferson")

# If the sheet is empty, set the column headers and alignment
if sheet.max_row == 1:
    sheet.append(["Page", "Keyword", "Title", "URL", "Text"])
    for cell in sheet[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Process each page
for page in range(1, num_pages + 1):

    # Construct the URL for the current page
    url = base_url + str(page)

    # Send a GET request to the URL
    response = requests.get(url)

    # Create a BeautifulSoup object to parse the HTML content
    soup = BeautifulSoup(response.content, "html.parser")

    # Find all the text within the <p> tags on the page
    text = ""
    for paragraph in soup.find_all("p"):
        text += paragraph.get_text() + "\n"

    # Process the text with spaCy
    doc = nlp(text)

    # Collect matching paragraphs
    matching_paragraphs = []
    for sent in doc.sents:
        # Check if any keyword matches the lemma of any word in the sentence
        if any(token.lemma_ in keywords for token in sent):
            matching_paragraphs.append(sent.text.strip())

    # Find the <h1> tag with class "title"
    title_tag = soup.find("h1", class_="title")
    title = title_tag.text if title_tag is not None else ''

    # Add matching references to the worksheet
    for paragraph in matching_paragraphs:
        paragraph_doc = nlp(paragraph)
        matched_keywords = [keyword for keyword in keywords if any(token.lemma_ == keyword for token in paragraph_doc)]
        sheet.append([page, ', '.join(matched_keywords), title, url, paragraph])
        reference_count += 1

        # Print the reference
        print(f"Reference - Page: {page}")
        print(f"Keyword: {', '.join(matched_keywords)}")
        print(f"Title: {title}")
        print(f"URL: {url}")
        print(f"Text: {paragraph}\n")

# Save the workbook
workbook.save(workbook_path)

# Print total reference count
print(f"Total References: {reference_count}")